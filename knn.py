#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Feb  8 17:39:07 2022

@author: enzobergamini
"""

import numpy as np
from PIL import Image
import openpyxl as x
import os
import matplotlib.pyplot as plt
import colorsys
from datetime import datetime


url_bdd="BDD/bdd3.xlsx"
n=7


def bdd(url_bdd):
    fichier = x.load_workbook(url_bdd)
    tableau=fichier.active
    LSM=[]
    LM=[]
    i=2
    while (tableau.cell(i,1).value!=None and tableau.cell(i,2)!=None):
        if tableau.cell(i,3).value == 1:
            LM.append([tableau.cell(i,1).value, tableau.cell(i,2).value])
        else:
            LSM.append([tableau.cell(i,1).value, tableau.cell(i,2).value])
        i=1+i
    return LM, LSM

def bdd2(url_bdd):
    fichier = x.load_workbook(url_bdd)
    tableau=fichier.active
    LSM=[]
    LM=[]
    i=2
    while (tableau.cell(i,1).value!=None and tableau.cell(i,2)!=None):
        if tableau.cell(i,4).value == 1:
            LM.append([tableau.cell(i,1).value, tableau.cell(i,2).value, tableau.cell(i,3).value])
        else:
            LSM.append([tableau.cell(i,1).value, tableau.cell(i,2).value, tableau.cell(i,3).value])
        i=1+i
    return LM, LSM


L1,L2=bdd2("BDD/bdd3.xlsx")

def graph1(L1,L2,U,d):
    plt.figure(dpi=500)
    plt.xlabel("pixel de teinte peau (% de l'image)")
    plt.ylabel("pixel de teinte bleu (% de l'image)")
    X1=[ L1[i][0] for i in range(len(L1))]
    Y1=[ L1[i][1] for i in range(len(L1))]
    plt.scatter(X1,Y1,1,marker="^", c="#2596be",label="visage avec masque")
    X2=[ L2[i][0] for i in range(len(L2))]
    Y2=[ L2[i][1] for i in range(len(L2))]
    plt.scatter(X2,Y2,1,marker="s", c="#D8832d",label="visage sans masque")
    plt.scatter(U[0],U[1],20,c="#A34638",label="visage à déterminer")
    plt.legend()
    plt.savefig("points12.png", dpi=500)
    plt.show()
    
def graph2(L1,L2,U,d):
    plt.figure(dpi=500)
    plt.xlabel("ecart type gris")
    plt.ylabel("ecart type teinte")
    X1=[ L1[i][0] for i in range(len(L1))]
    Y1=[ L1[i][1] for i in range(len(L1))]
    plt.scatter(X1,Y1,1,marker="^", c="#2596be")
    X2=[ L2[i][0] for i in range(len(L2))]
    Y2=[ L2[i][1] for i in range(len(L2))]
    plt.scatter(X2,Y2,1,marker="s", c="#D8832d")
    plt.scatter(U[0],U[1],20,c="#A34638")
    plt.savefig("points1.png", dpi=500)
    plt.show()
   
def graph3(L1,L2,U,d):
    plt.figure(dpi=500)
    plt.xlabel("ecart type tot")
    plt.ylabel("ecart type teinte")
    X1=[ L1[i][0] for i in range(len(L1))]
    Y1=[ L1[i][1] for i in range(len(L1))]
    plt.scatter(X1,Y1,1,marker="^", c="#2596be")
    X2=[ L2[i][0] for i in range(len(L2))]
    Y2=[ L2[i][1] for i in range(len(L2))]
    plt.scatter(X2,Y2,1,marker="s", c="#D8832d")
    plt.scatter(U[0],U[1],20,c="#A34638")
    plt.savefig("points2.png", dpi=500)
    plt.show()
    
def graph4(L1,L2,U,d):
    fig = plt.figure(dpi=500)
    ax = fig.add_subplot(projection='3d')
    ax.set_xlabel("ecart type teinte global")
    ax.set_ylabel("ecart type teinte local")
    ax.set_zlabel("ecart type nuance de gris local")
    X1=[ L1[i][0] for i in range(len(L1))]
    Y1=[ L1[i][1] for i in range(len(L1))]
    Z1=[ L1[i][2] for i in range(len(L1))]
    ax.scatter(X1,Y1,Z1)
    
    X2=[ L2[i][0] for i in range(len(L2))]
    Y2=[ L2[i][1] for i in range(len(L2))]
    Z2=[ L2[i][2] for i in range(len(L2))]
    ax.scatter(X2,Y2,Z2)
    ax.scatter(U[0],U[1],U[2],c="#A34638",s=30)
    ax.scatter(200,25,10,c="#A34638",s=30)
    ax.legend(("visage avec masque","visage sans masque","visage à déterminer"))
    
    
    plt.savefig(str((datetime.now()))+".png", dpi=500)
    plt.show()    


def distance_E(U,V):
    if len(U)!=len(V):
        return False
    n=len(U)
    s=0
    for i in range(n):
        s=s+(U[i]-V[i])**2
    return np.sqrt(s)

def knn(url_bdd, url_img ,n):
    LM,LSM = bdd(url_bdd)
    img=importation(url_img)
    Y=traitement(img)
    U=[Y[1], Y[0]]
    L1=[]
    for i in range(len(LM)):
        L1.append([distance_E(LM[i],U),1])
    L1=sorted(L1)
    L2=[]
    for i in range(len(LSM)):
        L2.append([distance_E(LSM[i],U),0])
    L2=sorted(L2)
    L=L1+L2
    L=sorted(L)
    s=0
    for i in range(n):
        print(L[i])
        s=s+L[i][1]
    graph1(LM, LSM, U, L[n])
    return (s/n)*100

def knn1(url_bdd, url_img ,n):
    LM,LSM = bdd(url_bdd)
    U=[ecart1(url_img), ecart2(url_img)]
    L1=[]
    for i in range(len(LM)):
        L1.append([distance_E(LM[i],U),1])
    L1=sorted(L1)
    L2=[]
    for i in range(len(LSM)):
        L2.append([distance_E(LSM[i],U),0])
    L2=sorted(L2)
    L=L1+L2
    L=sorted(L)
    s=0
    for i in range(n):
        print(L[i])
        s=s+L[i][1]
    graph2(LM, LSM, U, L[n])
    return (s/n)*100 ,"% masque"
 
def knn2(url_bdd, url_img ,n):
    LM,LSM = bdd(url_bdd)
    U=[ecart3(url_img), ecart2(url_img)]
    L1=[]
    for i in range(len(LM)):
        L1.append([distance_E(LM[i],U),1])
    L1=sorted(L1)
    L2=[]
    for i in range(len(LSM)):
        L2.append([distance_E(LSM[i],U),0])
    L2=sorted(L2)
    L=L1+L2
    L=sorted(L)
    s=0
    for i in range(n):
        print(L[i])
        s=s+L[i][1]
    graph3(LM, LSM, U, L[n])
    return (s/n)*100 ,"% masque"  


def knn3(url_bdd, url_img ,n):
    LM,LSM = bdd(url_bdd)
    U=[ecart3(url_img), ecart2(url_img)]
    L1=[]
    for i in range(len(LM)):
        L1.append([distance_E(LM[i],U),1])
    L1=sorted(L1)
    L2=[]
    for i in range(len(LSM)):
        L2.append([distance_E(LSM[i],U),0])
    L2=sorted(L2)
    L=L1+L2
    L=sorted(L)
    s=0
    for i in range(n):
        print(L[i])
        s=s+L[i][1]
    graph3(LM, LSM, U, L[n])
    return (s/n)*100 ,"% masque"  

def knn4(url_bdd, url_img ,n):
    LM,LSM = bdd2(url_bdd)
    U=[ecart3(url_img), ecart2(url_img), ecart1(url_img)]
    L1=[]
    for i in range(len(LM)):
        L1.append([distance_E(LM[i],U),1])
    L1=sorted(L1)
    L2=[]
    for i in range(len(LSM)):
        L2.append([distance_E(LSM[i],U),0])
    L2=sorted(L2)
    L=L1+L2
    L=sorted(L)
    s=0
    for i in range(n):
        print(L[i])
        s=s+L[i][1]
    graph4(LM, LSM, U, L[n])
    return (s/n)*100 ,"% masque"  
         
    
def importation(url):
    image=Image.open(url)
    return(np.array(image))    

def traitement(M):
    Z=np.shape(M)
    sb=0
    sp=0
    pixel=[0,0,0]
    for i in range(Z[0]):
        for j in range(Z[1]):
            pixel[0]+=M[i][j][0]
            pixel[1]+=M[i][j][1]
            pixel[2]+=M[i][j][2]
            if(int(convertHLSpixel(M[i][j][0],M[i][j][1],M[i][j][2]))>=120 and int(convertHLSpixel(M[i][j][0],M[i][j][1],M[i][j][2]))<=220):
                    sb+=1
            if(int(convertHLSpixel(M[i][j][0],M[i][j][1],M[i][j][2]))>=10 and int(convertHLSpixel(M[i][j][0],M[i][j][1],M[i][j][2]))<=25):
                    sp+=1
    return ((sb/(Z[0]*Z[1]))*100,(sp/(Z[0]*Z[1]))*100)

def ecart1(url):
    M=Image.open(url)
    M=M.convert("L")
    M=np.array(M)
    Z=M.shape
    img2=M[int(Z[0]*0.6):int(Z[0]*0.9), int(Z[1]*0.3):int(Z[1]*0.7)]
    Z=img2.shape  
    moy=0
    for i in range(Z[0]):
        for j in range(Z[1]):
            moy=moy+img2[i][j]
    moy=int(moy/(Z[0]*Z[1]))
    s=0
    for i in range(Z[0]):
        for j in range(Z[1]):
            s=s+(moy-img2[i][j])**2
    s=int(np.sqrt(s/(Z[0]*Z[1])))
    
    return s


def ecart2(url):
    M=importation(url)
    Z=M.shape
    img2=M[int(Z[0]*0.6):int(Z[0]*0.9), int(Z[1]*0.3):int(Z[1]*0.7)]
    Z=img2.shape  
    img=np.ones((Z[0],Z[1],1))
    moy=0
    for i in range(Z[0]):
        for j in range(Z[1]):
            moy=moy+colorsys.rgb_to_hls(img2[i][j][0]/256,img2[i][j][1]/256,img2[i][j][2]/256)[0]*360
            img[i][j]=colorsys.rgb_to_hls(img2[i][j][0]/256,img2[i][j][1]/256,img2[i][j][2]/256)[0]*360
    moy=int(moy/(Z[0]*Z[1]))
    s=0
    for i in range(Z[0]):
        for j in range(Z[1]):
            s=s+(moy-img[i][j])**2
    s=int(np.sqrt(s/(Z[0]*Z[1])))
    
    return s

def ecart3(url):
    M=importation(url)
    Z=M.shape
    img2=M[int(Z[0]*0.6):int(Z[0]*0.9), int(Z[1]*0.3):int(Z[1]*0.7)]
    Z=img2.shape  
    img=np.ones((Z[0],Z[1],1))
    moy=0
    for i in range(Z[0]):
        for j in range(Z[1]):
            moy=moy+colorsys.rgb_to_hls(img2[i][j][0]/256,img2[i][j][1]/256,img2[i][j][2]/256)[0]*360
            img[i][j]=colorsys.rgb_to_hls(img2[i][j][0]/256,img2[i][j][1]/256,img2[i][j][2]/256)[0]*360
    moy=int(moy/(Z[0]*Z[1]))
    Z=M.shape
    s=0
    for i in range(Z[0]):
        for j in range(Z[1]):
            s=s+(moy-colorsys.rgb_to_hls(M[i][j][0]/256,M[i][j][1]/256,M[i][j][2]/256)[0]*360)**2
    s=int(np.sqrt(s/(Z[0]*Z[1])))
    
    return s



def convertHLSpixel(r,g,b):
    somme=r+g+b
    r=r/255
    g=g/255
    b=b/255
    Max=max(r,g,b)
    Min=min(r,g,b)
    C=Max-Min
    L=(Max+Min)/2
    if L!=0:  
        S=C/(L*2)
    else:
        return 0
    if C!=0:
        if Max==r:
            T=(g-b)/C
        if Max==g:
            T=2.0+(b-r)/C
        if Max==b:
            T=4.0+(r-g)/C
    else:
        return 0
    return T*60